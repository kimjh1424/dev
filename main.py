import sys
import time
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException, ElementClickInterceptedException, StaleElementReferenceException
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
import re
from urllib.parse import quote
import os
import json
from datetime import datetime
import random
import shutil
import tempfile

class CrawlerThread(threading.Thread):
    def __init__(self, keyword, max_count, callback, status_callback, headless_mode=False, checkpoint_enabled=True, resume_from_checkpoint=None):
        super().__init__()
        self.keyword = keyword
        self.max_count = max_count
        self.callback = callback
        self.status_callback = status_callback
        self.is_running = True
        self.is_paused = False
        self.daemon = True
        self.headless_mode = headless_mode
        self.checkpoint_enabled = checkpoint_enabled
        self.resume_from_checkpoint = resume_from_checkpoint
        self.checkpoint_dir = f"checkpoints_{keyword}_{datetime.now().strftime('%Y%m%d_%H%M%S')}"
        self.session_count = 0  # 세션 카운터 추가
        self.manual_checkpoint_flag = False  # 수동 체크포인트 플래그 추가
        
        # 체크포인트 디렉토리 생성
        if checkpoint_enabled and not resume_from_checkpoint:
            os.makedirs(self.checkpoint_dir, exist_ok=True)

    def create_stealth_driver(self):
        """봇 감지 회피를 위한 강화된 드라이버 설정"""
        options = webdriver.ChromeOptions()
        
        # 헤드리스 모드 설정
        if self.headless_mode:
            options.add_argument("--headless")
            self.status_callback("헤드리스 모드 활성화")
        
        # 봇 감지 회피 설정
        options.add_argument("--disable-blink-features=AutomationControlled")
        options.add_experimental_option("excludeSwitches", ["enable-automation"])
        options.add_experimental_option('useAutomationExtension', False)
        
        # User-Agent 랜덤화
        user_agents = [
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/120.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/119.0.0.0 Safari/537.36',
            'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/118.0.0.0 Safari/537.36'
        ]
        options.add_argument(f'user-agent={random.choice(user_agents)}')
        
        # 프로필 격리 - 매 세션마다 새로운 임시 프로필 사용
        temp_profile = tempfile.mkdtemp()
        options.add_argument(f'--user-data-dir={temp_profile}')
        
        # 기타 안정성 옵션
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")
        options.add_argument("--disable-features=VizDisplayCompositor")
        options.add_argument("--disable-images")  # 이미지 로드 안함
        options.add_argument("--log-level=3")
        
        # 메모리 관리
        options.add_argument("--max_old_space_size=4096")
        options.add_argument("--memory-pressure-off")
        
        # 창 크기 랜덤화
        window_sizes = [(1920, 1080), (1366, 768), (1440, 900)]
        width, height = random.choice(window_sizes)
        options.add_argument(f'--window-size={width},{height}')
        
        options.page_load_strategy = 'eager'
        
        driver = webdriver.Chrome(options=options)
        driver.set_page_load_timeout(30)
        driver.implicitly_wait(5)
        
        # JavaScript로 webdriver 속성 숨기기
        driver.execute_script("""
            Object.defineProperty(navigator, 'webdriver', {
                get: () => undefined
            });
            Object.defineProperty(navigator, 'plugins', {
                get: () => [1, 2, 3, 4, 5]
            });
            Object.defineProperty(navigator, 'languages', {
                get: () => ['ko-KR', 'ko', 'en-US', 'en']
            });
        """)
        
        # 임시 프로필 경로 저장 (나중에 삭제용)
        driver.temp_profile = temp_profile
        
        return driver

    def find_next_page_button(self, driver):
        """다음 페이지 버튼을 정확히 찾는 함수"""
        selectors = ["a.eUTV2", "a._2PoiJ", "a[class*='page']", "button[class*='next']"]
        
        for selector in selectors:
            try:
                buttons = driver.find_elements(By.CSS_SELECTOR, selector)
                
                for button in buttons:
                    if button.get_attribute("aria-disabled") == "true":
                        continue
                    
                    try:
                        svg_element = button.find_element(By.TAG_NAME, "svg")
                        path_element = svg_element.find_element(By.TAG_NAME, "path")
                        d_attribute = path_element.get_attribute("d")
                        
                        if d_attribute and any(pattern in d_attribute for pattern in ["M12", "M14", "M10.524"]):
                            if not any(pattern in d_attribute for pattern in ["M8", "M6", "M13.476"]):
                                self.status_callback(f"다음 페이지 버튼 찾음: {selector}")
                                return button
                                
                    except:
                        button_text = button.text.strip()
                        if "다음" in button_text or ">" in button_text:
                            return button
            except:
                continue
                
        return None

    def turbo_scroll_to_load_all(self, driver, scroll_container):
        """터보 스크롤 - 확실하게 70개 모두 로드"""
        self.status_callback("⚡ 스크롤 컨테이너 찾는 중...")
        
        # 네이버 지도 전용 스크롤 컨테이너 찾기
        scroll_found = False
        scroll_container = None
        
        # 방법 1: ID로 직접 찾기
        try:
            scroll_container = driver.find_element(By.ID, "_list_scroll_container")
            self.status_callback("✅ _list_scroll_container 찾음!")
            scroll_found = True
        except:
            pass
        
        # 방법 2: 클래스로 찾기
        if not scroll_found:
            try:
                scroll_container = driver.find_element(By.CSS_SELECTOR, "div._2ky45")
                self.status_callback("✅ div._2ky45 찾음!")
                scroll_found = True
            except:
                pass
        
        # 방법 3: JavaScript로 스크롤 가능한 요소 찾기
        if not scroll_found:
            try:
                scroll_container = driver.execute_script("""
                    var divs = document.querySelectorAll('div');
                    for (var i = 0; i < divs.length; i++) {
                        if (divs[i].scrollHeight > divs[i].clientHeight && 
                            divs[i].scrollHeight > 500) {
                            return divs[i];
                        }
                    }
                    return null;
                """)
                if scroll_container:
                    self.status_callback("✅ JavaScript로 스크롤 컨테이너 찾음!")
                    scroll_found = True
            except:
                pass
        
        # 방법 4: iframe 내부 body
        if not scroll_found:
            try:
                scroll_container = driver.find_element(By.TAG_NAME, "body")
                self.status_callback("⚠️ body를 스크롤 컨테이너로 사용")
            except:
                self.status_callback("❌ 스크롤 컨테이너를 찾을 수 없음!")
                return 0
        
        # 초기 아이템 수 확인
        initial_items = driver.find_elements(By.CSS_SELECTOR, "a.place_bluelink")
        if not initial_items:
            initial_items = driver.find_elements(By.CSS_SELECTOR, "li.UEzoS")
        
        previous_count = len(initial_items)
        self.status_callback(f"초기 아이템: {previous_count}개")
        
        # 메인 스크롤 전략 - 끝까지 스크롤을 반복
        self.status_callback("⚡ 강화된 스크롤 시작...")
        no_change_count = 0
        max_no_change = 5  # 5번 연속 변화 없으면 종료
        
        for i in range(20):  # 최대 20번 시도로 증가
            try:
                # 1. JavaScript로 끝까지 스크롤
                driver.execute_script("""
                    var element = arguments[0];
                    element.scrollTop = element.scrollHeight;
                """, scroll_container)
                
                time.sleep(0.8)  # 로드 대기 시간 증가
                
                # 2. 추가로 조금씩 더 스크롤 (남은 요소 로드)
                for _ in range(3):
                    driver.execute_script("""
                        var element = arguments[0];
                        element.scrollTop = element.scrollTop + 500;
                    """, scroll_container)
                    time.sleep(0.3)
                
                # 현재 아이템 수 확인
                current_items = driver.find_elements(By.CSS_SELECTOR, "a.place_bluelink")
                if not current_items:
                    current_items = driver.find_elements(By.CSS_SELECTOR, "li.UEzoS")
                
                current_count = len(current_items)
                
                if current_count > previous_count:
                    self.status_callback(f"⚡ {current_count}개 로드됨 (+{current_count - previous_count})")
                    previous_count = current_count
                    no_change_count = 0  # 변화가 있으면 카운트 리셋
                    
                    # 70개 도달시에도 추가 스크롤
                    if current_count >= 70:
                        self.status_callback(f"✅ 70개 도달! 추가 확인 중...")
                        # 70개 도달 후에도 5번 더 스크롤해서 확인
                        for j in range(5):
                            driver.execute_script("""
                                var element = arguments[0];
                                element.scrollTop = element.scrollHeight;
                            """, scroll_container)
                            time.sleep(0.5)
                            
                            # 더 로드되는지 확인
                            extra_items = driver.find_elements(By.CSS_SELECTOR, "a.place_bluelink")
                            if not extra_items:
                                extra_items = driver.find_elements(By.CSS_SELECTOR, "li.UEzoS")
                            
                            if len(extra_items) > current_count:
                                current_count = len(extra_items)
                                self.status_callback(f"⚡ 추가 발견! 총 {current_count}개")
                        break
                else:
                    no_change_count += 1
                    if no_change_count >= max_no_change:
                        self.status_callback(f"⚠️ {max_no_change}번 연속 변화 없음. 최종: {current_count}개")
                        break
                    
            except Exception as e:
                self.status_callback(f"스크롤 오류: {e}")
        
        # 추가 스크롤 전략: 중간 위치들로 스크롤
        if current_count < 70:
            self.status_callback("🔄 추가 스크롤 전략 실행...")
            try:
                total_height = driver.execute_script("return arguments[0].scrollHeight", scroll_container)
                positions = [0.2, 0.4, 0.6, 0.8, 1.0]  # 20%, 40%, 60%, 80%, 100% 위치
                
                for pos in positions:
                    driver.execute_script(f"""
                        arguments[0].scrollTop = {int(total_height * pos)};
                    """, scroll_container)
                    time.sleep(0.5)
                    
                    # 아이템 수 다시 확인
                    current_items = driver.find_elements(By.CSS_SELECTOR, "a.place_bluelink")
                    if not current_items:
                        current_items = driver.find_elements(By.CSS_SELECTOR, "li.UEzoS")
                    
                    new_count = len(current_items)
                    if new_count > current_count:
                        current_count = new_count
                        self.status_callback(f"⚡ 추가 스크롤로 {current_count}개 발견!")
                        
            except:
                pass
        
        # 맨 위로 스크롤
        try:
            driver.execute_script("arguments[0].scrollTop = 0", scroll_container)
        except:
            driver.execute_script("window.scrollTo(0, 0);")
        
        time.sleep(0.5)
        
        # 최종 아이템 수 확인
        final_items = driver.find_elements(By.CSS_SELECTOR, "a.place_bluelink")
        if not final_items:
            final_items = driver.find_elements(By.CSS_SELECTOR, "li.UEzoS")
        
        final_count = len(final_items)
        self.status_callback(f"⚡ 스크롤 완료! 최종 로드: {final_count}개")
        
        if final_count < 70 and final_count > 20:
            self.status_callback(f"⚠️ 70개 미만 로드됨. 해당 검색어의 결과가 {final_count}개일 수 있습니다.")
        
        return final_count

    def save_checkpoint(self, data, checkpoint_num, page_num, item_index):
        """체크포인트 저장"""
        if not self.checkpoint_enabled:
            return
            
        checkpoint_data = {
            'keyword': self.keyword,
            'checkpoint_num': checkpoint_num,
            'total_collected': len(data),
            'current_page': page_num,
            'current_item_index': item_index,
            'timestamp': datetime.now().isoformat(),
            'data': data
        }
        
        # JSON 파일로 상태 저장
        checkpoint_file = os.path.join(self.checkpoint_dir, f'checkpoint_{checkpoint_num}.json')
        with open(checkpoint_file, 'w', encoding='utf-8') as f:
            json.dump(checkpoint_data, f, ensure_ascii=False, indent=2)
        
        # Excel 파일로도 저장
        excel_file = os.path.join(self.checkpoint_dir, f'checkpoint_{checkpoint_num}_{len(data)}개.xlsx')
        self.save_to_excel_internal(data, excel_file)
        
        self.status_callback(f"💾 체크포인트 {checkpoint_num} 저장 완료! ({len(data)}개)")

    def save_to_excel_internal(self, data, file_path):
        """내부용 엑셀 저장 함수"""
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "크롤링 결과"
        
        # 헤더
        headers = ["번호", "장소명", "도로명 주소", "지번 주소", "전화번호"]
        sheet.append(headers)
        
        # 헤더 스타일
        for cell in sheet[1]:
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
            cell.fill = openpyxl.styles.PatternFill(
                start_color="FF6B00", 
                end_color="FF6B00", 
                fill_type="solid"
            )
        
        # 데이터 추가
        for idx, item in enumerate(data, 1):
            sheet.append([idx] + item)
        
        # 열 너비 조정
        for column in sheet.columns:
            max_length = 0
            for cell in column:
                try:
                    if cell.value and len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            
            if hasattr(column[0], 'column_letter'):
                adjusted_width = min(max_length + 2, 50)
                sheet.column_dimensions[column[0].column_letter].width = adjusted_width
        
        workbook.save(file_path)

    def run(self):
        data = []
        collected_count = 0
        current_page = 1
        checkpoint_num = 0
        
        # 체크포인트에서 재개하는 경우
        if self.resume_from_checkpoint:
            self.status_callback(f"🔄 체크포인트에서 재개 중...")
            try:
                with open(self.resume_from_checkpoint, 'r', encoding='utf-8') as f:
                    checkpoint_data = json.load(f)
                    data = checkpoint_data['data']
                    collected_count = len(data)
                    current_page = checkpoint_data['current_page']
                    checkpoint_num = checkpoint_data['checkpoint_num']
                    self.checkpoint_dir = os.path.dirname(self.resume_from_checkpoint)
                    self.status_callback(f"✅ 체크포인트 로드 완료! ({collected_count}개 기존 데이터)")
            except Exception as e:
                self.status_callback(f"❌ 체크포인트 로드 실패: {e}")
                return
        
        self.status_callback(f"🚀 '{self.keyword}' 크롤링 시작! (목표: {self.max_count}개)")
        self.status_callback("💡 체크포인트 모드 - 100개마다 자동 저장")
        if self.headless_mode:
            self.status_callback("👻 헤드리스 모드로 실행중...")
        else:
            self.status_callback("👀 일반 모드로 실행중 (브라우저 표시)")
        self.status_callback("=" * 50)
        
        driver = None
        search_url = None

        try:
            # 100개 단위로 세션 분할
            while collected_count < self.max_count and self.is_running:
                # 일시정지 확인
                while self.is_paused and self.is_running:
                    time.sleep(0.5)
                
                if not self.is_running:
                    break
                
                # 세션당 최대 수집 개수 (100개 또는 남은 개수)
                session_max = min(100, self.max_count - collected_count)
                session_collected = 0
                
                self.session_count += 1
                self.status_callback(f"\n{'='*50}")
                self.status_callback(f"🔄 세션 {self.session_count} 시작 (목표: {session_max}개)")
                self.status_callback(f"🔧 봇 감지 회피를 위한 새 브라우저 프로필 생성...")
                self.status_callback(f"{'='*50}")
                
                # 세션 간 대기 시간 (봇 감지 회피)
                if self.session_count > 1:
                    wait_time = random.randint(5, 10)
                    self.status_callback(f"⏳ 네이버 봇 감지 회피를 위해 {wait_time}초 대기...")
                    for i in range(wait_time):
                        if not self.is_running:
                            break
                        time.sleep(1)
                        self.status_callback(f"⏳ 대기 중... {wait_time - i - 1}초")
                
                # 강화된 드라이버 생성
                driver = self.create_stealth_driver()
                
                encoded_keyword = quote(self.keyword)
                search_url = f"https://map.naver.com/p/search/{encoded_keyword}"
                self.status_callback(f"검색 URL: {search_url}")
                
                # 페이지 로드 전 대기 (랜덤)
                time.sleep(random.uniform(2, 4))
                
                driver.get(search_url)
                time.sleep(random.uniform(3, 5))  # 초기 로드 대기 랜덤화

                self.status_callback("searchIframe으로 전환...")
                WebDriverWait(driver, 10).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "searchIframe")))
                time.sleep(random.uniform(1, 2))
                self.status_callback("✅ searchIframe 전환 성공")
                
                # 초기 페이지 로드 확인
                try:
                    WebDriverWait(driver, 10).until(
                        EC.presence_of_element_located((By.CSS_SELECTOR, "a.place_bluelink, li.UEzoS"))
                    )
                    self.status_callback("✅ 검색 결과 로드 확인")
                except:
                    self.status_callback("❌ 검색 결과를 찾을 수 없습니다.")
                    if driver:
                        driver.quit()
                        if hasattr(driver, 'temp_profile'):
                            shutil.rmtree(driver.temp_profile, ignore_errors=True)
                    continue

                # 지정된 페이지로 이동 (재개하는 경우)
                if current_page > 1 and self.resume_from_checkpoint:
                    self.status_callback(f"📌 {current_page} 페이지로 이동 중...")
                    for _ in range(current_page - 1):
                        next_button = self.find_next_page_button(driver)
                        if next_button:
                            driver.execute_script("arguments[0].click();", next_button)
                            time.sleep(2)
                        else:
                            break
                    self.resume_from_checkpoint = None  # 한 번만 실행

                # 세션 내에서 크롤링
                while session_collected < session_max and self.is_running:
                    # 일시정지 확인
                    while self.is_paused and self.is_running:
                        time.sleep(0.5)
                    
                    if not self.is_running:
                        break
                    
                    self.status_callback(f"\n{'='*50}")
                    self.status_callback(f"📌 {current_page} 페이지 크롤링 시작")
                    self.status_callback(f"📌 현재까지 수집: {collected_count}개 (세션 {self.session_count}: {session_collected}개)")
                    self.status_callback(f"{'='*50}")

                    # 터보 스크롤로 모든 아이템 로드
                    loaded_count = self.turbo_scroll_to_load_all(driver, None)
                    
                    # 모든 장소 요소 가져오기
                    all_place_elements = []
                    place_selectors = ["a.place_bluelink", "li.UEzoS.rTjJo", "li.UEzoS"]
                    
                    for selector in place_selectors:
                        try:
                            all_place_elements = driver.find_elements(By.CSS_SELECTOR, selector)
                            if all_place_elements:
                                self.status_callback(f"장소 요소 찾음: {selector} ({len(all_place_elements)}개)")
                                break
                        except:
                            continue

                    if not all_place_elements:
                        self.status_callback(f"❌ {current_page} 페이지에서 장소를 찾을 수 없습니다.")
                        break

                    page_target_count = len(all_place_elements)
                    page_collected = 0
                    self.status_callback(f"⚡ {page_target_count}개 장소 크롤링 시작!")

                    # 각 장소 크롤링
                    for i, element in enumerate(all_place_elements):
                        if not self.is_running or session_collected >= session_max:
                            break
                        
                        # 일시정지 확인
                        while self.is_paused and self.is_running:
                            time.sleep(0.5)

                        try:
                            # 요소 클릭
                            if element.tag_name == 'a':
                                element_to_click = element
                            else:
                                try:
                                    element_to_click = element.find_element(By.CSS_SELECTOR, "a.place_bluelink")
                                except:
                                    element_to_click = element.find_element(By.CSS_SELECTOR, "a")

                            # JavaScript 클릭 (더 빠름)
                            driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element_to_click)
                            time.sleep(random.uniform(0.3, 0.6))  # 랜덤 대기
                            driver.execute_script("arguments[0].click();", element_to_click)
                            time.sleep(random.uniform(1, 1.5))  # 랜덤 대기

                            driver.switch_to.default_content()
                            WebDriverWait(driver, 5).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "entryIframe")))
                            time.sleep(0.5)

                            # 데이터 추출
                            place_data = driver.execute_script("""
                                var result = {
                                    name: '',
                                    roadAddr: '', 
                                    jibunAddr: '',
                                    phone: ''
                                };
                                
                                // 이름
                                var nameElem = document.querySelector('.YwYLL, ._3Apjo, .GHAhO, h2');
                                if (nameElem) result.name = nameElem.textContent.replace('복사', '').trim();
                                
                                // 도로명 주소
                                var roadElem = document.querySelector('.LDgIH');
                                if (roadElem) result.roadAddr = roadElem.textContent.replace('복사', '').trim();
                                
                                // 전화번호
                                var phoneElem = document.querySelector('.xlx7Q, ._3ZA58 span, .dry01, .J7eF_');
                                if (phoneElem) {
                                    result.phone = phoneElem.textContent
                                        .replace('휴대전화번호', '')
                                        .replace('복사', '')
                                        .trim();
                                }
                                
                                return result;
                            """)

                            name = place_data.get('name', '정보 없음')
                            road_address = place_data.get('roadAddr', '정보 없음')
                            jibun_address = "정보 없음"
                            phone = place_data.get('phone', '정보 없음')
                            
                            # 지번주소 가져오기
                            try:
                                address_button = driver.find_element(By.CSS_SELECTOR, "a.PkgBl")
                                driver.execute_script("arguments[0].click();", address_button)
                                time.sleep(0.3)
                                
                                # 지번주소 찾기
                                jibun_found = False
                                
                                try:
                                    time.sleep(0.2)
                                    address_items = driver.find_elements(By.CSS_SELECTOR, ".nQ7Lh")
                                    for item in address_items:
                                        item_text = item.text.strip()
                                        if "지번" in item_text:
                                            jibun_address = item_text.replace("지번", "").replace("복사", "").strip()
                                            jibun_found = True
                                            break
                                except:
                                    pass
                                
                                # 버튼 다시 클릭해서 닫기
                                if jibun_found:
                                    try:
                                        close_button = driver.find_element(By.CSS_SELECTOR, "a.PkgBl[aria-expanded='true']")
                                        driver.execute_script("arguments[0].click();", close_button)
                                    except:
                                        pass
                                    
                            except:
                                pass

                            # 전화번호 버튼 클릭 시도
                            if phone == "정보 없음" or not any(char.isdigit() for char in phone):
                                max_retries = 5
                                retry_count = 0
                                
                                while retry_count < max_retries and (phone == "정보 없음" or not any(char.isdigit() for char in phone)):
                                    try:
                                        phone_button = driver.find_element(By.CSS_SELECTOR, "a.BfF3H")
                                        if "전화번호 보기" in phone_button.text or "전화번호" in phone_button.text:
                                            driver.execute_script("arguments[0].click();", phone_button)
                                            time.sleep(0.5)
                                            
                                            phone_selectors = [".J7eF_", ".xlx7Q", "._3ZA58 span", ".dry01"]
                                            
                                            for selector in phone_selectors:
                                                try:
                                                    phone_elem = driver.find_element(By.CSS_SELECTOR, selector)
                                                    temp_phone = phone_elem.text.replace("휴대전화번호", "").replace("복사", "").replace("안내", "").strip()
                                                    
                                                    if temp_phone and sum(c.isdigit() for c in temp_phone) >= 7:
                                                        phone = temp_phone
                                                        break
                                                except:
                                                    continue
                                            
                                            if phone != "정보 없음" and any(char.isdigit() for char in phone):
                                                break
                                                
                                    except:
                                        pass
                                    
                                    retry_count += 1

                            # 데이터 저장
                            if name != "정보 없음":
                                data.append([name, road_address, jibun_address, phone])
                                collected_count += 1
                                session_collected += 1
                                page_collected += 1
                                
                                # 진행 상황 로그
                                if collected_count % 10 == 0:
                                    self.status_callback(f"✅ ({collected_count}/{self.max_count}) 수집 진행중...")
                                
                                # 100개마다 자동 체크포인트
                                if self.checkpoint_enabled and collected_count % 100 == 0:
                                    checkpoint_num += 1
                                    self.save_checkpoint(data, checkpoint_num, current_page, i)
                                
                                # 수동 체크포인트 확인
                                if self.manual_checkpoint_flag:
                                    checkpoint_num += 1
                                    self.save_checkpoint(data, checkpoint_num, current_page, i)
                                    self.manual_checkpoint_flag = False
                                    self.status_callback("💾 수동 체크포인트 저장 완료!")
                                
                                # 목표 달성 확인
                                if collected_count >= self.max_count:
                                    break
                            else:
                                page_collected += 1

                        except Exception as e:
                            page_collected += 1
                            continue
                        finally:
                            try:
                                driver.switch_to.default_content()
                                WebDriverWait(driver, 5).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "searchIframe")))
                                time.sleep(0.3)
                            except:
                                pass

                    self.status_callback(f"⚡ {current_page} 페이지 완료! ({page_collected}/{page_target_count}개 수집)")
                    
                    # 세션 목표 달성 확인
                    if session_collected >= session_max:
                        break
                    
                    # 다음 페이지로 이동
                    if page_collected >= page_target_count * 0.8:  # 80% 이상 수집시 다음 페이지
                        try:
                            driver.switch_to.default_content()
                            WebDriverWait(driver, 5).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "searchIframe")))
                            
                            next_button = self.find_next_page_button(driver)
                            if next_button:
                                driver.execute_script("arguments[0].click();", next_button)
                                self.status_callback(f"⚡ 다음 페이지로 이동!")
                                current_page += 1
                                time.sleep(2)
                            else:
                                self.status_callback("❌ 마지막 페이지입니다.")
                                break
                                
                        except Exception as e:
                            self.status_callback(f"❌ 페이지 이동 실패: {e}")
                            break
                    elif page_target_count <= 20:
                        # 페이지에 20개 이하만 있는 경우 다음 페이지로
                        self.status_callback(f"📌 이 페이지는 {page_target_count}개만 있습니다. 다음 페이지로 이동합니다.")
                        try:
                            driver.switch_to.default_content()
                            WebDriverWait(driver, 5).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "searchIframe")))
                            
                            next_button = self.find_next_page_button(driver)
                            if next_button:
                                driver.execute_script("arguments[0].click();", next_button)
                                current_page += 1
                                time.sleep(2)
                            else:
                                self.status_callback("❌ 마지막 페이지입니다.")
                                break
                        except:
                            break
                    else:
                        self.status_callback(f"❌ 수집률이 낮습니다. ({page_collected}/{page_target_count})")
                        self.status_callback("🔍 스크롤이 충분하지 않았을 수 있습니다. 다음 세션에서 재시도합니다.")
                        # 낮은 수집률이어도 계속 진행 (다음 세션에서 재시도)
                        break
                
                # 드라이버 종료 및 프로필 삭제
                if driver:
                    driver.quit()
                    if hasattr(driver, 'temp_profile'):
                        shutil.rmtree(driver.temp_profile, ignore_errors=True)
                        self.status_callback("🧹 임시 브라우저 프로필 삭제 완료")
                    driver = None
                
                # 세션 완료 메시지
                self.status_callback(f"\n✅ 세션 {self.session_count} 완료! ({session_collected}개 수집)")
                
            self.status_callback(f"\n🎉 크롤링 완료!\n총 {collected_count}개 수집")

        except Exception as e:
            self.status_callback(f"❌ 크롤링 오류: {str(e)[:100]}...")
        finally:
            if driver:
                driver.quit()
                if hasattr(driver, 'temp_profile'):
                    shutil.rmtree(driver.temp_profile, ignore_errors=True)
            
            # 마지막 체크포인트 저장
            if self.checkpoint_enabled and len(data) > 0:
                if len(data) % 100 != 0:  # 100개 단위가 아닌 경우에만
                    checkpoint_num += 1
                    self.save_checkpoint(data, checkpoint_num, current_page, 0)
            
            self.status_callback(f"최종 수집 데이터: {len(data)}개")
            if self.root:
                self.root.after(0, self.callback, data)
            else:
                self.callback(data)

    def stop(self):
        self.is_running = False
        
    def pause(self):
        self.is_paused = True
        self.status_callback("⏸️ 크롤링 일시정지")
        
    def resume(self):
        self.is_paused = False
        self.status_callback("▶️ 크롤링 재개")


class NaverMapCrawlerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Naver Map Crawler v5.0 - 체크포인트 버전")
        self.root.geometry("900x750")
        
        # 스타일 설정
        style = ttk.Style()
        style.configure('Title.TLabel', font=('Arial', 18, 'bold'))
        style.configure('Turbo.TLabel', font=('Arial', 12, 'bold'), foreground='#ff6b00')
        
        self.crawler_thread = None
        self.setup_ui()
        
    def setup_ui(self):
        # 메인 프레임
        main_frame = ttk.Frame(self.root, padding="10")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 윈도우 크기 조정
        self.root.grid_rowconfigure(0, weight=1)
        self.root.grid_columnconfigure(0, weight=1)
        main_frame.grid_rowconfigure(7, weight=1)
        
        # 타이틀
        title_label = ttk.Label(main_frame, text="🚀 네이버 지도 크롤러 v5.0", style='Title.TLabel')
        title_label.grid(row=0, column=0, columnspan=3, pady=(0, 5))
        
        # 부제목
        subtitle_label = ttk.Label(main_frame, text="체크포인트 자동 저장 지원", style='Turbo.TLabel')
        subtitle_label.grid(row=1, column=0, columnspan=3, pady=(0, 10))
        
        # 검색 영역
        search_frame = ttk.Frame(main_frame)
        search_frame.grid(row=2, column=0, columnspan=3, pady=5)
        
        ttk.Label(search_frame, text="검색어:").grid(row=0, column=0, padx=5)
        self.search_entry = ttk.Entry(search_frame, width=30)
        self.search_entry.grid(row=0, column=1, padx=5)
        self.search_entry.bind('<Return>', lambda e: self.start_crawling())
        
        ttk.Label(search_frame, text="최대 갯수:").grid(row=0, column=2, padx=5)
        self.max_count_var = tk.StringVar(value="300")
        self.max_count_spinbox = ttk.Spinbox(search_frame, from_=1, to=1000, textvariable=self.max_count_var, width=10)
        self.max_count_spinbox.grid(row=0, column=3, padx=5)
        
        self.search_button = ttk.Button(search_frame, text="🚀 크롤링 시작", command=self.start_crawling)
        self.search_button.grid(row=0, column=4, padx=10)
        
        # 크롤링 제어 버튼들
        control_frame = ttk.Frame(main_frame)
        control_frame.grid(row=3, column=0, columnspan=3, pady=5)
        
        self.pause_button = ttk.Button(control_frame, text="⏸️ 일시정지", command=self.pause_crawling, state='disabled')
        self.pause_button.grid(row=0, column=0, padx=5)
        
        self.resume_button = ttk.Button(control_frame, text="▶️ 재개", command=self.resume_crawling, state='disabled')
        self.resume_button.grid(row=0, column=1, padx=5)
        
        self.stop_button = ttk.Button(control_frame, text="⏹️ 중지", command=self.stop_crawling, state='disabled')
        self.stop_button.grid(row=0, column=2, padx=5)
        
        self.checkpoint_button = ttk.Button(control_frame, text="💾 수동 체크포인트", command=self.manual_checkpoint, state='disabled')
        self.checkpoint_button.grid(row=0, column=3, padx=5)
        
        self.load_checkpoint_button = ttk.Button(control_frame, text="📂 체크포인트 불러오기", command=self.load_checkpoint)
        self.load_checkpoint_button.grid(row=0, column=4, padx=5)
        
        # 옵션 프레임
        option_frame = ttk.LabelFrame(main_frame, text="⚙️ 크롤링 옵션", padding="5")
        option_frame.grid(row=4, column=0, columnspan=3, pady=10, sticky=(tk.W, tk.E))
        
        # 헤드리스 모드 체크박스
        self.headless_var = tk.BooleanVar(value=False)
        self.headless_checkbox = ttk.Checkbutton(
            option_frame,
            text="👻 헤드리스 모드 (브라우저 숨김 - 메모리 절약)",
            variable=self.headless_var
        )
        self.headless_checkbox.grid(row=0, column=0, padx=10, pady=5)
        
        # 체크포인트 자동 저장 체크박스
        self.checkpoint_var = tk.BooleanVar(value=True)
        self.checkpoint_checkbox = ttk.Checkbutton(
            option_frame,
            text="💾 100개마다 자동 체크포인트 저장",
            variable=self.checkpoint_var
        )
        self.checkpoint_checkbox.grid(row=0, column=1, padx=10, pady=5)
        
        # 모드 설명
        mode_info = tk.Label(
            option_frame, 
            text="💡 안내사항\n• 100개마다 자동으로 체크포인트가 생성됩니다\n• 네이버 봇 감지 회피를 위해 100개마다 새 브라우저 세션을 시작합니다\n• 체크포인트에서 이어서 크롤링이 가능합니다\n• 수동 체크포인트도 언제든 저장 가능합니다",
            justify=tk.LEFT,
            fg='blue'
        )
        mode_info.grid(row=1, column=0, columnspan=2, padx=10, pady=5)
        
        # 안내 메시지
        info_text = "• 체크포인트 자동 저장으로 안정적인 대량 수집\n• 네이버 봇 감지 회피 강화\n• 일시정지/재개 기능 지원"
        info_label = ttk.Label(main_frame, text=info_text, foreground='gray')
        info_label.grid(row=6, column=0, columnspan=3, pady=5)
        
        # 로그 프레임
        log_frame = ttk.LabelFrame(main_frame, text="📋 크롤링 로그", padding="5")
        log_frame.grid(row=7, column=0, columnspan=3, sticky=(tk.W, tk.E, tk.N, tk.S), pady=10)
        log_frame.grid_rowconfigure(0, weight=1)
        log_frame.grid_columnconfigure(0, weight=1)
        
        # 로그 텍스트 위젯
        self.log_text = tk.Text(
            log_frame, 
            height=15, 
            wrap=tk.WORD,
            bg="#1a1a1a",
            fg="#00ff88",
            insertbackground="#00ff88",
            selectbackground="#0078d4",
            selectforeground="#ffffff",
            font=('Consolas', 10)
        )
        self.log_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 스크롤바
        scrollbar = ttk.Scrollbar(log_frame, orient=tk.VERTICAL, command=self.log_text.yview)
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        self.log_text.configure(yscrollcommand=scrollbar.set)
        
        # 로그 색상 태그 설정
        self.log_text.tag_configure("info", foreground="#00ff88")
        self.log_text.tag_configure("success", foreground="#4ec9b0")
        self.log_text.tag_configure("warning", foreground="#ffcc00")
        self.log_text.tag_configure("error", foreground="#ff6b6b")
        self.log_text.tag_configure("turbo", foreground="#ff6b00", font=('Consolas', 10, 'bold'))
        
        # 상태바
        self.status_var = tk.StringVar(value="🚀 준비 완료")
        self.status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN)
        self.status_bar.grid(row=8, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(5, 0))
        
        # By 라벨
        by_label = ttk.Label(main_frame, text="By ANYCODER | v5.0 - 체크포인트", foreground='#ff6b00')
        by_label.grid(row=9, column=0, columnspan=3, pady=(5, 0))

    def start_crawling(self, resume_checkpoint=None):
        keyword = self.search_entry.get().strip()
        if not keyword:
            messagebox.showwarning("입력 오류", "검색어를 입력해주세요.")
            return
        
        try:
            max_count = int(self.max_count_var.get())
            if max_count < 1:
                raise ValueError
        except:
            messagebox.showwarning("입력 오류", "올바른 숫자를 입력해주세요.")
            return
            
        # 버튼 상태 변경
        self.search_button.config(state='disabled', text="🚀 크롤링 진행중...")
        self.search_entry.config(state='disabled')
        self.max_count_spinbox.config(state='disabled')
        self.headless_checkbox.config(state='disabled')
        self.checkpoint_checkbox.config(state='disabled')
        self.pause_button.config(state='normal')
        self.stop_button.config(state='normal')
        self.checkpoint_button.config(state='normal')
        self.load_checkpoint_button.config(state='disabled')
        
        # 로그 초기화
        if not resume_checkpoint:
            self.log_text.delete(1.0, tk.END)
        
        # 크롤링 스레드 시작
        headless_mode = self.headless_var.get()
        checkpoint_enabled = self.checkpoint_var.get()
        
        self.crawler_thread = CrawlerThread(
            keyword, 
            max_count, 
            self.crawling_finished, 
            self.update_status,
            headless_mode,
            checkpoint_enabled,
            resume_checkpoint
        )
        self.crawler_thread.root = self.root
        self.crawler_thread.start()
        
    def pause_crawling(self):
        if self.crawler_thread and self.crawler_thread.is_alive():
            self.crawler_thread.pause()
            self.pause_button.config(state='disabled')
            self.resume_button.config(state='normal')
            
    def resume_crawling(self):
        if self.crawler_thread and self.crawler_thread.is_alive():
            self.crawler_thread.resume()
            self.pause_button.config(state='normal')
            self.resume_button.config(state='disabled')
            
    def stop_crawling(self):
        if self.crawler_thread and self.crawler_thread.is_alive():
            result = messagebox.askquestion(
                "크롤링 중지",
                "크롤링을 중지하시겠습니까?\n현재까지 수집된 데이터는 저장됩니다."
            )
            if result == 'yes':
                self.crawler_thread.stop()
                self.status_var.set("⏹️ 크롤링 중지됨")
                
    def manual_checkpoint(self):
        if self.crawler_thread and self.crawler_thread.is_alive():
            try:
                # 크롤러 스레드에 수동 체크포인트 요청
                if hasattr(self.crawler_thread, 'manual_checkpoint_flag'):
                    self.crawler_thread.manual_checkpoint_flag = True
                    self.status_var.set("💾 수동 체크포인트 저장 요청...")
                    messagebox.showinfo("수동 체크포인트", "다음 장소 수집 후 체크포인트가 저장됩니다.")
            except Exception as e:
                messagebox.showerror("오류", f"체크포인트 저장 실패: {str(e)}")
            
    def load_checkpoint(self):
        checkpoint_file = filedialog.askopenfilename(
            title="체크포인트 파일 선택",
            filetypes=[("JSON files", "*.json"), ("All files", "*.*")]
        )
        
        if checkpoint_file:
            try:
                with open(checkpoint_file, 'r', encoding='utf-8') as f:
                    checkpoint_data = json.load(f)
                    
                # 체크포인트 정보 표시
                info_msg = f"체크포인트 정보:\n"
                info_msg += f"검색어: {checkpoint_data['keyword']}\n"
                info_msg += f"수집된 데이터: {checkpoint_data['total_collected']}개\n"
                info_msg += f"현재 페이지: {checkpoint_data['current_page']}\n"
                info_msg += f"저장 시간: {checkpoint_data['timestamp']}\n\n"
                info_msg += "이어서 크롤링하시겠습니까?"
                
                result = messagebox.askyesno("체크포인트 불러오기", info_msg)
                
                if result:
                    # 검색어 설정
                    self.search_entry.delete(0, tk.END)
                    self.search_entry.insert(0, checkpoint_data['keyword'])
                    
                    # 크롤링 시작
                    self.start_crawling(resume_checkpoint=checkpoint_file)
                    
            except Exception as e:
                messagebox.showerror("오류", f"체크포인트 파일을 불러올 수 없습니다:\n{str(e)}")
        
    def update_status(self, message):
        self.status_var.set(message)
        
        # 로그창에 메시지 추가
        timestamp = time.strftime("%H:%M:%S")
        log_message = f"[{timestamp}] {message}\n"
        
        # 메시지 타입에 따라 색상 적용
        tag = "info"
        if "⚡" in message or "🚀" in message:
            tag = "turbo"
        elif "✅" in message or "완료" in message:
            tag = "success"
        elif "경고" in message or "⚠️" in message:
            tag = "warning"
        elif "오류" in message or "실패" in message or "❌" in message:
            tag = "error"
        
        # 로그 텍스트에 추가
        self.log_text.insert(tk.END, log_message, tag)
        self.log_text.see(tk.END)
        self.root.update_idletasks()
        
    def crawling_finished(self, data):
        # 버튼 상태 복원
        self.search_button.config(state='normal', text="🚀 크롤링 시작")
        self.search_entry.config(state='normal')
        self.max_count_spinbox.config(state='normal')
        self.headless_checkbox.config(state='normal')
        self.checkpoint_checkbox.config(state='normal')
        self.pause_button.config(state='disabled')
        self.resume_button.config(state='disabled')
        self.stop_button.config(state='disabled')
        self.checkpoint_button.config(state='disabled')
        self.load_checkpoint_button.config(state='normal')
        
        if data:
            self.status_var.set(f"🚀 크롤링 완료! {len(data)}개 수집!")
            self.save_to_excel(data)
        else:
            messagebox.showinfo("결과 없음", "수집된 데이터가 없습니다.\n검색어를 확인하고 다시 시도해주세요.")
            self.status_var.set("데이터 수집 실패")
            
    def save_to_excel(self, data):
        if not data:
            return
            
        keyword = self.search_entry.get().strip()
        default_filename = f"네이버지도_{keyword}_{time.strftime('%Y%m%d_%H%M%S')}_최종.xlsx"
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=default_filename
        )
        
        if file_path:
            try:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "크롤링 결과"
                
                # 헤더
                headers = ["번호", "장소명", "도로명 주소", "지번 주소", "전화번호"]
                sheet.append(headers)
                
                # 헤더 스타일
                for cell in sheet[1]:
                    cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color="FF6B00", 
                        end_color="FF6B00", 
                        fill_type="solid"
                    )
                
                # 데이터 추가
                for idx, item in enumerate(data, 1):
                    sheet.append([idx] + item)
                
                # 열 너비 조정
                for column in sheet.columns:
                    max_length = 0
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    if hasattr(column[0], 'column_letter'):
                        adjusted_width = min(max_length + 2, 50)
                        sheet.column_dimensions[column[0].column_letter].width = adjusted_width
                    
                workbook.save(file_path)
                
                # 체크포인트 디렉토리 정보도 표시
                checkpoint_info = ""
                if self.crawler_thread and hasattr(self.crawler_thread, 'checkpoint_dir'):
                    checkpoint_info = f"\n체크포인트 저장 위치: {self.crawler_thread.checkpoint_dir}"
                
                messagebox.showinfo(
                    "저장 완료",
                    f"🚀 저장 완료!\n\n"
                    f"파일: {file_path}\n"
                    f"수집 데이터: {len(data)}개"
                    f"{checkpoint_info}"
                )
                self.status_var.set(f"✅ 저장 완료: {file_path}")
                
            except Exception as e:
                messagebox.showerror("저장 실패", f"엑셀 파일 저장 중 오류가 발생했습니다:\n{str(e)}")
                self.status_var.set("저장 실패")


def main():
    root = tk.Tk()
    app = NaverMapCrawlerApp(root)
    
    # 종료 시 크롤링 중지
    def on_closing():
        if app.crawler_thread and app.crawler_thread.is_alive():
            result = messagebox.askquestion(
                "작업 중",
                "크롤링 작업이 아직 진행 중입니다.\n종료하시겠습니까?"
            )
            if result == 'yes':
                app.crawler_thread.stop()
                root.destroy()
        else:
            root.destroy()
    
    root.protocol("WM_DELETE_WINDOW", on_closing)
    root.mainloop()


if __name__ == '__main__':
    main()
