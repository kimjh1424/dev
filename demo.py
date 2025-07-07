import sys
import time
import tkinter as tk
from tkinter import ttk, messagebox, filedialog
import threading
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, NoSuchElementException
from selenium.webdriver.common.action_chains import ActionChains
import openpyxl
import re
from urllib.parse import quote


class CrawlerThread(threading.Thread):
    def __init__(self, keyword, max_count, callback, status_callback):
        super().__init__()
        self.keyword = keyword
        self.max_count = 3  # 데모 버전은 3개 고정
        self.callback = callback
        self.status_callback = status_callback
        self.is_running = True
        self.daemon = True

    def run(self):
        data = []
        self.status_callback(f"[데모 버전] '{self.keyword}' 검색을 시작합니다... (최대 3개만 수집)")
        driver = None
        
        try:
            options = webdriver.ChromeOptions()
            options.add_argument("--log-level=3")
            options.add_argument("--disable-blink-features=AutomationControlled")
            options.add_experimental_option("excludeSwitches", ["enable-automation"])
            options.add_experimental_option('useAutomationExtension', False)
            driver = webdriver.Chrome(options=options)
            
            # URL 기반으로 직접 검색
            encoded_keyword = quote(self.keyword)
            search_url = f"https://map.naver.com/p/search/{encoded_keyword}"
            driver.get(search_url)
            
            time.sleep(3)

            # searchIframe으로 전환
            WebDriverWait(driver, 20).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "searchIframe")))
            time.sleep(2)

            # 스크롤 최소화 (데모 버전)
            self.status_callback("[데모 버전] 검색 결과를 불러오는 중...")
            time.sleep(2)

            # 장소 링크 찾기
            place_elements = []
            selectors = [
                "a.place_bluelink",
                "a[class*='place_bluelink']",
                ".place_bluelink",
                "span.YwYLL",
                "a[role='button']",
                ".VLTHu.OW9LQ"
            ]
            
            for selector in selectors:
                try:
                    elements = driver.find_elements(By.CSS_SELECTOR, selector)
                    if elements:
                        self.status_callback(f"[데모 버전] {len(elements)}개의 요소를 찾았습니다.")
                        if selector == "span.YwYLL":
                            place_elements = []
                            for elem in elements:
                                try:
                                    parent = elem.find_element(By.XPATH, "./ancestor::a[@role='button']")
                                    place_elements.append(parent)
                                except:
                                    try:
                                        parent = elem.find_element(By.XPATH, "./ancestor::a")
                                        place_elements.append(parent)
                                    except:
                                        pass
                        else:
                            place_elements = elements
                        break
                except NoSuchElementException:
                    continue
            
            if not place_elements:
                self.status_callback("[데모 버전] 클릭 가능한 장소를 찾을 수 없습니다.")
                self.callback(data)
                return

            # 최대 3개만 처리
            total_to_process = min(len(place_elements), 3)
            self.status_callback(f"[데모 버전] 3개의 정보만 수집합니다.")

            # 각 장소 클릭하여 정보 추출
            collected_count = 0
            for i in range(total_to_process):
                if not self.is_running:
                    break
                
                try:
                    # 매번 요소 다시 찾기
                    driver.switch_to.default_content()
                    WebDriverWait(driver, 10).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "searchIframe")))
                    time.sleep(1)
                    
                    # 현재 인덱스의 요소 다시 찾기
                    current_elements = []
                    for selector in selectors:
                        try:
                            elements = driver.find_elements(By.CSS_SELECTOR, selector)
                            if elements and len(elements) > i:
                                if selector == "span.YwYLL":
                                    try:
                                        parent = elements[i].find_element(By.XPATH, "./ancestor::a[@role='button']")
                                        current_elements = [parent]
                                    except:
                                        try:
                                            parent = elements[i].find_element(By.XPATH, "./ancestor::a")
                                            current_elements = [parent]
                                        except:
                                            continue
                                else:
                                    current_elements = [elements[i]]
                                break
                        except:
                            continue
                    
                    if not current_elements:
                        continue
                    
                    element = current_elements[0]
                    
                    # 스크롤
                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", element)
                    time.sleep(1)
                    
                    # 클릭
                    try:
                        driver.execute_script("arguments[0].click();", element)
                    except:
                        try:
                            element.click()
                        except:
                            continue
                    
                    time.sleep(2)
                    
                    # entryIframe으로 전환
                    driver.switch_to.default_content()
                    
                    try:
                        WebDriverWait(driver, 5).until(EC.frame_to_be_available_and_switch_to_it((By.ID, "entryIframe")))
                        time.sleep(1)
                        
                        # 정보 추출
                        name = "정보 없음"
                        road_address = "정보 없음"
                        jibun_address = "정보 없음"
                        phone = "정보 없음"
                        
                        # 이름 추출
                        name_selectors = [".YwYLL", ".GHAhO", "span.YwYLL", "h2.YwYLL"]
                        for sel in name_selectors:
                            try:
                                name_elem = WebDriverWait(driver, 3).until(
                                    EC.presence_of_element_located((By.CSS_SELECTOR, sel))
                                )
                                name = name_elem.text.strip()
                                name = name.replace('복사', '').strip()
                                if name:
                                    break
                            except:
                                continue
                        
                        # 주소 추출
                        try:
                            address_button = driver.find_element(By.CSS_SELECTOR, "a.PkgBl")
                            driver.execute_script("arguments[0].click();", address_button)
                            time.sleep(1)
                            
                            address_divs = driver.find_elements(By.CSS_SELECTOR, "div.nQ7Lh")
                            for div in address_divs:
                                try:
                                    address_text = div.text.strip()
                                    address_text = address_text.replace('복사', '').replace('도로명', '').replace('지번', '').strip()
                                    
                                    if ('로' in address_text or '길' in address_text) and road_address == "정보 없음":
                                        road_address = address_text
                                    elif ('동' in address_text or '리' in address_text or re.search(r'\d+-\d+', address_text)) and jibun_address == "정보 없음":
                                        jibun_address = address_text
                                except:
                                    continue
                        except:
                            pass
                        
                        # 전화번호 추출
                        try:
                            phone_button_selectors = [".BfF3H", ".U7pYf", "button[aria-label*='전화']"]
                            for sel in phone_button_selectors:
                                try:
                                    phone_button = driver.find_element(By.CSS_SELECTOR, sel)
                                    driver.execute_script("arguments[0].click();", phone_button)
                                    time.sleep(1)
                                    break
                                except:
                                    continue
                            
                            phone_selectors = [".J7eF_", ".xlx7Q", ".RiCN3", "span.xlx7Q", "a[href^='tel:']"]
                            for sel in phone_selectors:
                                try:
                                    phone_elem = driver.find_element(By.CSS_SELECTOR, sel)
                                    phone_text = phone_elem.text.strip()
                                    if phone_text and re.search(r'\d{2,}', phone_text):
                                        phone_text = phone_text.replace('복사', '').replace('휴대전화번호', '').replace('전화번호', '').replace('전화', '').strip()
                                        phone = ' '.join(phone_text.split())
                                        break
                                except:
                                    continue
                        except:
                            pass
                        
                        if name != "정보 없음":
                            # 최종 정제
                            if road_address != "정보 없음":
                                road_address = road_address.replace('도로명', '').replace('복사', '').replace('주소', '').strip()
                                road_address = road_address.strip('[](){}|').strip()
                            if jibun_address != "정보 없음":
                                jibun_address = jibun_address.replace('지번', '').replace('복사', '').replace('주소', '').strip()
                                jibun_address = jibun_address.strip('[](){}|').strip()
                            if phone != "정보 없음":
                                phone = phone.replace('휴대전화번호', '').replace('전화번호', '').replace('전화', '').replace('복사', '').replace('연락처', '').strip()
                                phone = phone.strip('[](){}|').strip()
                                phone = ' '.join(phone.split())
                            
                            data.append([name, road_address, jibun_address, phone])
                            collected_count += 1
                            self.status_callback(f"[데모] ({collected_count}/3) {name} 정보 수집 완료")
                            
                            if collected_count >= 3:
                                self.status_callback(f"[데모 버전] 3개 수집 완료. 더 많은 정보는 정식 버전에서!")
                                break
                        
                    except TimeoutException:
                        self.status_callback(f"[데모] ({i + 1}/3번째 항목에서 상세 정보를 찾을 수 없음)")
                    except Exception as e:
                        self.status_callback(f"[데모] ({i + 1}/3번째 항목 처리 중 오류: {str(e)}")
                    
                except Exception as e:
                    self.status_callback(f"[데모] 항목 처리 중 오류: {str(e)}")
                    continue

            self.status_callback(f"[데모 버전] 크롤링 완료. 총 {collected_count}개의 정보를 수집했습니다.")

        except Exception as e:
            self.status_callback(f"[데모] 크롤링 프로세스 중 오류 발생: {str(e)}")
        finally:
            if driver:
                driver.quit()
            self.callback(data)

    def stop(self):
        self.is_running = False


class NaverMapCrawlerApp:
    def __init__(self, root):
        self.root = root
        self.root.title("Naver Map Crawler v2.0 - DEMO VERSION")
        self.root.geometry("700x400")
        
        # 스타일 설정
        style = ttk.Style()
        style.configure('Title.TLabel', font=('Arial', 18, 'bold'))
        style.configure('Demo.TLabel', font=('Arial', 12, 'bold'), foreground='red')
        
        self.crawler_thread = None
        self.setup_ui()
        
    def setup_ui(self):
        # 메인 프레임
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # 데모 버전 안내
        demo_label = ttk.Label(main_frame, text="🔒 데모 버전 - 최대 3개 항목만 크롤링 가능", style='Demo.TLabel')
        demo_label.grid(row=0, column=0, columnspan=3, pady=(0, 10))
        
        # 타이틀
        title_label = ttk.Label(main_frame, text="네이버 지도 크롤러", style='Title.TLabel')
        title_label.grid(row=1, column=0, columnspan=3, pady=(0, 20))
        
        # 검색 영역
        search_frame = ttk.Frame(main_frame)
        search_frame.grid(row=2, column=0, columnspan=3, pady=10)
        
        ttk.Label(search_frame, text="검색어:").grid(row=0, column=0, padx=5)
        self.search_entry = ttk.Entry(search_frame, width=40)
        self.search_entry.grid(row=0, column=1, padx=5)
        self.search_entry.bind('<Return>', lambda e: self.start_crawling())
        
        ttk.Label(search_frame, text="최대 갯수:").grid(row=0, column=2, padx=5)
        self.max_count_var = tk.StringVar(value="3")
        max_count_label = ttk.Label(search_frame, text="3 (고정)", foreground='gray')
        max_count_label.grid(row=0, column=3, padx=5)
        
        self.search_button = ttk.Button(search_frame, text="검색 시작", command=self.start_crawling)
        self.search_button.grid(row=0, column=4, padx=10)
        
        # 안내 메시지
        info_text = "• 검색 결과를 엑셀 파일로 저장합니다.\n• 수집 항목: 장소명, 도로명 주소, 지번 주소, 전화번호\n• ⚠️ 데모 버전은 3개 항목만 수집 가능합니다."
        info_label = ttk.Label(main_frame, text=info_text, foreground='gray')
        info_label.grid(row=3, column=0, columnspan=3, pady=20)
        
        # 정식 버전 안내
        purchase_button = ttk.Button(main_frame, text="🚀 정식 버전 구매하기 (최대 500개 수집 가능)", 
                                   command=self.show_purchase_info)
        purchase_button.grid(row=4, column=0, columnspan=3, pady=10)
        
        # 상태바
        self.status_var = tk.StringVar(value="[데모 버전] 준비 완료")
        self.status_bar = ttk.Label(main_frame, textvariable=self.status_var, relief=tk.SUNKEN)
        self.status_bar.grid(row=5, column=0, columnspan=3, sticky=(tk.W, tk.E), pady=(20, 0))
        
        # By 라벨
        by_label = ttk.Label(main_frame, text="By ANYCODER | v2.0 DEMO", foreground='gray')
        by_label.grid(row=6, column=0, columnspan=3, pady=(10, 0))
        
    def show_purchase_info(self):
        messagebox.showinfo(
            "정식 버전 안내",
            "정식 버전 특징:\n\n"
            "✅ 최대 500개까지 데이터 수집 가능\n"
            "✅ 빠른 크롤링 속도\n"
            "✅ 무제한 사용\n"
            "✅ 기술 지원 제공\n\n"
            "문의: kjh1424@proton.me"
        )
        
    def start_crawling(self):
        keyword = self.search_entry.get().strip()
        if not keyword:
            messagebox.showwarning("입력 오류", "검색어를 입력해주세요.")
            return
            
        # 데모 버전 안내
        result = messagebox.askquestion(
            "데모 버전 안내",
            "데모 버전은 3개의 정보만 수집할 수 있습니다.\n계속하시겠습니까?"
        )
        
        if result == 'no':
            return
            
        self.search_button.config(state='disabled', text="크롤링 중...")
        self.search_entry.config(state='disabled')
        
        self.crawler_thread = CrawlerThread(keyword, 3, self.crawling_finished, self.update_status)
        self.crawler_thread.start()
        
    def update_status(self, message):
        self.status_var.set(message)
        self.root.update_idletasks()
        
    def crawling_finished(self, data):
        self.search_button.config(state='normal', text="검색 시작")
        self.search_entry.config(state='normal')
        
        if data:
            self.status_var.set(f"[데모 버전] 크롤링 완료. {len(data)}개의 정보를 수집했습니다.")
            self.save_to_excel(data)
        else:
            messagebox.showinfo("결과 없음", "수집된 데이터가 없습니다.\n검색어를 확인하고 다시 시도해주세요.")
            self.status_var.set("[데모] 데이터 수집 실패")
            
    def save_to_excel(self, data):
        if not data:
            return
            
        keyword = self.search_entry.get().strip()
        default_filename = f"네이버지도_DEMO_{keyword}_{time.strftime('%Y%m%d_%H%M%S')}.xlsx"
        
        file_path = filedialog.asksaveasfilename(
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile=default_filename
        )
        
        if file_path:
            try:
                workbook = openpyxl.Workbook()
                sheet = workbook.active
                sheet.title = "네이버 지도 크롤링 결과 (DEMO)"
                
                # 데모 버전 워터마크
                sheet['A1'] = "DEMO VERSION - 정식 버전은 최대 500개까지 수집 가능"
                sheet['A1'].font = openpyxl.styles.Font(bold=True, color="FF0000")
                sheet.merge_cells('A1:E1')
                
                # 헤더
                headers = ["번호", "장소명", "도로명 주소", "지번 주소", "전화번호"]
                sheet.append(headers)
                
                # 헤더 스타일
                for cell in sheet[2]:
                    cell.font = openpyxl.styles.Font(bold=True)
                    cell.fill = openpyxl.styles.PatternFill(
                        start_color="DDDDDD", 
                        end_color="DDDDDD", 
                        fill_type="solid"
                    )
                
                # 데이터 추가
                for idx, item in enumerate(data, 1):
                    sheet.append([idx] + item)
                
                # 열 너비 조정
                for column in sheet.columns:
                    max_length = 0
                    for cell in column:
                        try:
                            if cell.value and len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    if hasattr(column[0], 'column_letter'):
                        adjusted_width = min(max_length + 2, 50)
                        sheet.column_dimensions[column[0].column_letter].width = adjusted_width
                    
                workbook.save(file_path)
                
                messagebox.showinfo(
                    "저장 완료",
                    f"'{file_path}'에 성공적으로 저장되었습니다.\n\n"
                    f"수집된 데이터: {len(data)}개\n\n"
                    f"💡 더 많은 데이터가 필요하시면 정식 버전을 구매하세요!"
                )
                self.status_var.set(f"[데모] 저장 완료: {file_path}")
                
            except Exception as e:
                messagebox.showerror("저장 실패", f"엑셀 파일 저장 중 오류가 발생했습니다:\n{str(e)}")
                self.status_var.set("[데모] 저장 실패")


def main():
    root = tk.Tk()
    app = NaverMapCrawlerApp(root)
    root.mainloop()


if __name__ == '__main__':
    main()